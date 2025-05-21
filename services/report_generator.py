import json
import re
import asyncio
import pandas as pd
import httpx
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from functools import lru_cache
from pathlib import Path
from datetime import date, timedelta, datetime
from typing import Any, Dict, List
from aiogram.types import Message
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Report
from services.logging import logger


# ------------------ HTTP‑clients ------------------
SYNC_CLIENT  = httpx.Client(timeout=30.0)
ASYNC_CLIENT = httpx.AsyncClient(timeout=30.0)


async def run_with_progress(message: Message, title: str, coro, *args):
    """
    Отображает сообщение с прогрессом, пока выполняется coroutine coro.
    Каждую секунду обновляет текст сообщения с индикацией выполнения.
    После завершения работы coroutine сообщение удаляется, а результат возвращается.
    В случае если API WB долго не выдает отчет - завершает coro и выбрасывает RuntimeError.
    Так же RuntimeError выбрасывается в случае неверного токена.
    """
    progress_message = await message.answer(f'{title}')
    task = asyncio.create_task(coro(*args))
    progress_stages = ['.', '..', '...']
    i = 0
    try:
        while not task.done():
            stage = progress_stages[i % len(progress_stages)]
            try:
                await progress_message.edit_text(f"{title}{stage}")
            except Exception as e:
                logger.error(f"Ошибка обновления прогресса: {e}")
            await asyncio.sleep(1)
            i += 1
            if i > 480:
                logger.error('Canceling task, report generation timeout')
                task.cancel()
                try:
                    await task  # Ждем завершения отмены
                except asyncio.CancelledError:
                    raise RuntimeError(
                        'Сервера Wildberries не отвечают слишком долго, мы сожалеем, но это от нас не зависит\n'
                        'Попробуйте позже.\n\n'
                        'Количество Ваших оставшихся генераций отчетов осталось неизменным'
                    )


        result = await task
        await progress_message.delete()
        return result
    except httpx.HTTPStatusError as e:
        logger.error(f'Ошибка запроса: {e}')
        raise RuntimeError(
            'У вас неверно введен токен магазина, или не выданы все нужные разрешения!\n'
            'Пересоздайте магазин и сгенерируйте отчет заново\n\n'
            'Количество Ваших оставшихся генераций отчетов осталось неизменным'
        )
    except Exception as e:
        await progress_message.delete()
        raise e


def get_weeks_range(count):
    today = date.today()
    previous_monday = today - timedelta(days=today.weekday()) - timedelta(days=7)
    weeks_range = []
    for i in range(count):
        week_start = previous_monday - timedelta(weeks=i)
        week_end = week_start + timedelta(days=6)
        weeks_range.append(f'{week_start.strftime("%d.%m.%Y")}-{week_end.strftime("%d.%m.%Y")}')

    return weeks_range


async def orm_add_report(session: AsyncSession, tg_id: int, date_of_week: date, report_path: str, store_id: int):
    obj = Report(
        tg_id=tg_id,
        date_of_week=date_of_week,
        report_path=report_path,
        store_id=store_id,
    )
    session.add(obj)
    await session.commit()


def get_dates_from_str(dates):
    """transform dates DD.MM.YYYY-DD.MM.YYYY to YYYY-MM-DD, YYYY-MM-DD"""
    dates = dates.split('-')
    start_date, end_date = dates[0].split('.'), dates[1].split('.')
    start_date = f'{start_date[2]}-{start_date[1]}-{start_date[0]}'
    end_date = f'{end_date[2]}-{end_date[1]}-{end_date[0]}'
    return start_date, end_date


def get_dates_in_range(start: str, end: str) -> List[str]:
    s = datetime.strptime(start, "%Y-%m-%d").date()
    e = datetime.strptime(end,   "%Y-%m-%d").date()
    dates: List[str] = []
    while s <= e:
        dates.append(s.isoformat())
        s += timedelta(days=1)
    return dates


def create_empty_adv_report() -> pd.DataFrame:
    return pd.DataFrame(columns=["Артикул WB", "totalAdjustedSum", "Period"])


@lru_cache(maxsize=1)
def fetch_product_cards_mapping(token: str) -> Dict[str, Dict[str, str]]:
    logger.info("Загрузка маппинга карточек товара...")
    url = "https://content-api.wildberries.ru/content/v2/get/cards/list"
    headers = {"Authorization": token, "Content-Type": "application/json"}
    payload = {"settings": {"cursor": {"limit": 100}, "filter": {"withPhoto": -1}}}
    mapping: Dict[str, Dict[str, str]] = {}
    while True:
        resp = SYNC_CLIENT.post(url, headers=headers, json=payload)
        resp.raise_for_status()
        data = resp.json()
        cards = data.get("cards", [])
        for c in cards:
            key = str(c.get("nmID") or c.get("nmId", "")).strip().upper()
            mapping[key] = {
                "vendorCode": c.get("vendorCode", "").strip(),
                "nmName":     (c.get("name") or c.get("nmName") or "").strip()
            }
        cursor = data.get("cursor") or {}
        if not cursor or len(cards) < 100:
            break
        payload["settings"]["cursor"].update({
            "updatedAt": cursor.get("updatedAt"),
            "nmID":      cursor.get("nmID")
        })
    logger.info("Маппинг карточек загружен: %d записей", len(mapping))
    return mapping


# ------------------ Sales Report ------------------

async def fetch_sales_records_async(date_from: str, date_to: str, token: str) -> List[Dict[str, Any]]:
    logger.info("Начинаем загрузку отчёта по продажам с %s по %s", date_from, date_to)
    url = "https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod"
    headers = {"Authorization": token, "Content-Type": "application/json"}
    records, rrdid = [], 0
    while True:
        resp = await ASYNC_CLIENT.get(
            url, headers=headers,
            params={"dateFrom": date_from, "dateTo": date_to, "rrdid": rrdid, "limit": 100000}
        )
        resp.raise_for_status()
        chunk = resp.json()
        if not chunk:
            break
        records.extend(chunk)
        last = chunk[-1]
        new_rrdid = last.get("rrd_id") or last.get("rrdid")
        if not new_rrdid or new_rrdid == rrdid:
            break
        rrdid = new_rrdid
        await asyncio.sleep(0.1)
    logger.info("Загрузка отчёта по продажам завершена: %d записей", len(records))
    return records

def transform_sales_records(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=[
            "Артикул WB", "Короткое название товара",
            "SUM из Кол-во", "SUM из Сумма продаж", "SUM из К перечислению продавцу",
            "SUM из Кол-во доставок", "SUM из Стоимость доставки",
            "SUM из Штрафы", "SUM из Дополнительный платеж",
            "Утилизация", "Подписка «Джем»"
        ])
    df = df.copy()
    if "deduction" in df.columns:
        df["deduction"] = pd.to_numeric(df["deduction"], errors="coerce").fillna(0)
    if "bonusTypeName" in df.columns:
        df.rename(columns={"bonusTypeName": "bonus_type_name"}, inplace=True)
    total_util = df.loc[
        df["bonus_type_name"].str.contains("утилизации", case=False, na=False) & (df["deduction"] != 0),
        "deduction"
    ].sum()
    total_jam = df.loc[
        df["bonus_type_name"].str.contains("джем", case=False, na=False) & (df["deduction"] != 0),
        "deduction"
    ].sum()
    if "vendorCode" not in df.columns:
        df["vendorCode"] = df.get("sa_name", "")
    df["Короткое название товара"] = (
        df["vendorCode"].fillna("").astype(str).str.strip().str.upper()
    )
    df.loc[df["Короткое название товара"].isin(["", "NAN"]), "Короткое название товара"] = "НЕОПОЗНАННЫЙ ТОВАР"
    df = df[df["nm_id"] != 0]
    sales_df = df[df["doc_type_name"] != "Возврат"].copy()
    returns_df = df[df["doc_type_name"] == "Возврат"].copy()
    cols = ["nm_id", "Короткое название товара"]
    sales_agg = sales_df.groupby(cols, as_index=False).agg({
        "quantity":"sum","retail_amount":"sum","ppvz_for_pay":"sum",
        "delivery_amount":"sum","delivery_rub":"sum","penalty":"sum","additional_payment":"sum"
    })
    sales_agg.rename(columns={
        "quantity":"SUM из Кол-во","retail_amount":"SUM из Сумма продаж",
        "ppvz_for_pay":"SUM из К перечислению продавцу","delivery_amount":"SUM из Кол-во доставок",
        "delivery_rub":"SUM из Стоимость доставки","penalty":"SUM из Штрафы",
        "additional_payment":"SUM из Дополнительный платеж"
    }, inplace=True)
    cnt = len(sales_agg)
    sales_agg["Утилизация"] = round(total_util/cnt,2) if cnt else 0.0
    sales_agg["Подписка «Джем»"] = round(total_jam/cnt,2) if cnt else 0.0
    returns_agg = returns_df.groupby(cols, as_index=False).agg({
        "quantity":"sum","retail_amount":"sum","ppvz_for_pay":"sum"
    })
    returns_agg.rename(columns={
        "quantity":"Возвраты (Кол-во)","retail_amount":"Возвраты (Сумма продаж)",
        "ppvz_for_pay":"Возвраты (К перечислению продавцу)"
    }, inplace=True)
    merged = pd.merge(sales_agg, returns_agg, on=cols, how="left")
    for c in ["Возвраты (Кол-во)","Возвраты (Сумма продаж)","Возвраты (К перечислению продавцу)"]:
        merged[c] = merged[c].fillna(0)
    merged.rename(columns={"nm_id":"Артикул WB"}, inplace=True)
    return merged


# ------------------ Storage Report ------------------

async def get_storage_report(date_from: str, date_to: str, token: str) -> pd.DataFrame:
    logger.info("Запрос отчёта по платному хранению... %s – %s", date_from, date_to)
    base, headers = "https://seller-analytics-api.wildberries.ru/api/v1/paid_storage", {"Authorization":f"Bearer {token}"}
    # create
    for backoff in [5,10,20,40,80]:
        resp = await ASYNC_CLIENT.get(base, headers=headers, params={"dateFrom": date_from, "dateTo": date_to})
        if resp.status_code != 429:
            resp.raise_for_status()
            break
        logger.warning("429 при создании хранения, жду %s", backoff); await asyncio.sleep(backoff)
    else:
        return pd.DataFrame(columns=["nmId","nmName","vendorCode","totalStorageSum","Period"])
    task = resp.json()["data"]["taskId"]
    status_url = f"{base}/tasks/{task}/status"
    # poll
    for _ in range(12):
        st = await ASYNC_CLIENT.get(status_url, headers=headers)
        if st.status_code == 429:
            await asyncio.sleep(5)
            continue
        st.raise_for_status()
        if st.json()["data"]["status"].lower()=="done":
            break
        await asyncio.sleep(5)
    else:
        return pd.DataFrame(columns=["nmId","nmName","vendorCode","totalStorageSum","Period"])
    # download
    dl_url = f"{base}/tasks/{task}/download"
    for backoff in [5,10,20,40,80]:
        dl = await ASYNC_CLIENT.get(dl_url, headers=headers)
        if dl.status_code != 429:
            dl.raise_for_status()
            data = dl.json()
            break
        logger.warning("429 при скачивании хранения, жду %s", backoff); await asyncio.sleep(backoff)
    else:
        return pd.DataFrame(columns=["nmId","nmName","vendorCode","totalStorageSum","Period"])
    df = pd.DataFrame(data)
    df["Цена склада"] = pd.to_numeric(df.get("warehousePrice",0),errors="coerce").fillna(0)
    grp = df.groupby("nmId",as_index=False)["Цена склада"].sum().rename(columns={"Цена склада":"totalStorageSum"})
    grp["nmId"] = grp["nmId"].astype(str).str.upper()
    grp["totalStorageSum"] = grp["totalStorageSum"].round(2)
    cards = fetch_product_cards_mapping(token)
    grp = grp.assign(
        vendorCode=grp["nmId"].map(lambda x: cards.get(x,{}).get("vendorCode",x)),
        nmName    =grp["nmId"].map(lambda x: cards.get(x,{}).get("nmName",x).upper()),
        Period    =f"{date_from} - {date_to}"
    )
    logger.info("Отчёт по хранению готов: %d позиций", len(grp))
    return grp[["nmId","nmName","vendorCode","totalStorageSum","Period"]]


# ------------------ Acceptance report ------------------

async def get_acceptance_report(date_from: str, date_to: str, token: str) -> pd.DataFrame:
    logger.info("Запрос отчёта по платной приёмке... %s – %s", date_from, date_to)
    base, headers = "https://seller-analytics-api.wildberries.ru/api/v1/acceptance_report", {"Authorization":token}
    # create
    for backoff in [5,10,20,40,80]:
        resp = await ASYNC_CLIENT.get(base, headers=headers, params={"dateFrom": date_from, "dateTo": date_to})
        if resp.status_code != 429:
            resp.raise_for_status()
            break
        logger.warning("429 при создании приёмки, жду %s", backoff); await asyncio.sleep(backoff)
    else:
        return pd.DataFrame(columns=["Артикул WB","Платная приемка"])
    task = resp.json()["data"]["taskId"]
    status_url = f"{base}/tasks/{task}/status"
    while True:
        await asyncio.sleep(5)
        st = await ASYNC_CLIENT.get(status_url, headers=headers)
        st.raise_for_status()
        if st.json()["data"]["status"].lower()=="done":
            break
    dl = await ASYNC_CLIENT.get(f"{base}/tasks/{task}/download", headers=headers)
    dl.raise_for_status()
    data = dl.json()
    if not isinstance(data,list) or not data:
        return pd.DataFrame(columns=["Артикул WB","Платная приемка"])
    df_ac = pd.DataFrame(data)
    nm_col = next((c for c in df_ac.columns if re.search(r"(?i)nm[_]?id$",c)),None)
    if nm_col is None:
        nm_col = next((c for c in df_ac.columns if re.search(r"(?i)nm.*id",c)),None)
    if nm_col is None:
        return pd.DataFrame(columns=["Артикул WB","Платная приемка"])
    df_ac["Артикул WB"] = df_ac[nm_col].astype(str).str.upper()
    ac = df_ac.groupby("Артикул WB",as_index=False)["total"].sum().rename(columns={"total":"Платная приемка"})
    logger.info("Отчёт по приёмке готов: %d позиций",len(ac))
    return ac


# ------------------ Adds report ------------------

def get_ad_expenses_report(token: str, doc_number: str, period_end: str) -> pd.DataFrame:
    logger.info("Формирование отчёта по рекламе, updNum=%s", doc_number)
    if not doc_number:
        return create_empty_adv_report()
    elif ' ' in doc_number:
        upd = set([int(num) for num in doc_number.split()])
    else:
        upd = {int(doc_number)}

    end_date = datetime.strptime(period_end, "%Y-%m-%d").date()
    fr, to = (end_date - timedelta(days=30)).isoformat(), period_end
    period = f"{fr} - {to}"
    headers = {"Authorization": token}

    # Запрос списка рекламных документов
    upd_list = SYNC_CLIENT.get(
        f"https://advert-api.wildberries.ru/adv/v1/upd?from={fr}&to={to}",
        headers=headers
    )
    upd_list.raise_for_status()

    # Фильтрация по номерам документов
    items = [x for x in upd_list.json() if x.get("updNum") in upd]
    if not items:
        return create_empty_adv_report()

    # Суммируем расходы по кампаниям
    summary = {}
    for it in items:
        cid = it.get("advertId") or it.get("id")
        summary[cid] = summary.get(cid, 0.0) + float(it.get("updSum") or 0)

    # Формируем payload для каждого номера документа
    payload = []
    for doc_num in upd:
        payload.extend([
            {"id": cid, "updNum": doc_num, "dates": get_dates_in_range(fr, to)}
            for cid in summary
        ])

    # Запрос детальной статистики
    full = SYNC_CLIENT.post(
        "https://advert-api.wildberries.ru/adv/v2/fullstats",
        headers={**headers, "Content-Type": "application/json"},
        content=json.dumps(payload)
    )
    full.raise_for_status()

    data2 = full.json()
    if not isinstance(data2, list):
        return create_empty_adv_report()

    # Агрегация данных по товарам
    agg = {}
    for camp in data2:
        cid = camp.get("advertId") or camp.get("id")
        fact = summary.get(cid, 0.0)
        raw_total = sum(
            float(nm.get("sum") or 0)
            for day in camp.get("days", [])
            for app in day.get("apps", [])
            for nm in app.get("nm", [])
        ) or 0.0
        coef = fact / raw_total if raw_total > 0 else 1.0

        for day in camp.get("days", []):
            for app in day.get("apps", []):
                for nm in app.get("nm", []):
                    nid = nm.get("nmId")
                    name = nm.get("name") or ""
                    val = float(nm.get("sum") or 0) * coef
                    entry = agg.setdefault(nid, {"nmName": name, "totalAdjustedSum": 0.0})
                    entry["totalAdjustedSum"] += val
                    if not entry["nmName"] and name:
                        entry["nmName"] = name

    # Формируем DataFrame
    rows = [
        [str(nid).upper(), round(e["totalAdjustedSum"], 2), period, e["nmName"]]
        for nid, e in agg.items()
    ]
    df_adv = pd.DataFrame(
        rows,
        columns=["Артикул WB", "totalAdjustedSum", "Period", "Название товара"]
    )
    logger.info("Отчёт по рекламе готов: %d позиций", len(df_adv))
    return df_adv


# ------------------ Генерация отчёта ------------------

async def generate_report_with_params(dates: str, doc_number: str, store_token: str, store_name: str, tg_id: int, store_id: int) -> str:
    logger.info("Старт отчёта для %s: %s",store_name,dates)
    start_date, end_date = get_dates_from_str(dates)

    sales_task      = fetch_sales_records_async(f"{start_date}T00:00:00",f"{end_date}T23:59:59",store_token)
    acceptance_task = get_acceptance_report(start_date,end_date,store_token)
    storage_task    = get_storage_report(start_date,end_date,store_token)
    advert_task     = asyncio.to_thread(get_ad_expenses_report,store_token,doc_number,end_date)

    raw_records, acceptance_df, storage_df, adv_df = await asyncio.gather(
        sales_task, acceptance_task, storage_task, advert_task
    )

    df_raw   = pd.DataFrame(raw_records)
    sales_df = transform_sales_records(df_raw)

    # расширяем приёмку
    sa_sum=0.0
    if "deduction" in df_raw.columns and "bonusTypeName" in df_raw.columns:
        df_raw["deduction"]=pd.to_numeric(df_raw["deduction"],errors="coerce").fillna(0)
        df_raw["bonus_type_name"]=df_raw["bonusTypeName"].astype(str)
        sa_sum = df_raw.loc[
            df_raw["bonus_type_name"].str.contains("при[её]м",case=False,na=False)&(df_raw["deduction"]!=0),
            "deduction"
        ].sum()
    api_sum=acceptance_df["Платная приемка"].sum() if not acceptance_df.empty else 0.0
    if abs(api_sum-sa_sum)>1e-6:
        prev=(datetime.strptime(start_date,"%Y-%m-%d").date()-timedelta(days=2)).isoformat()
        try:
            acceptance_df=await get_acceptance_report(prev,end_date,store_token)
        except Exception as e:
            logger.warning("Не удалось расширить приёмку: %s",e)

    # отзывы и прочее
    reviews_agg=pd.DataFrame(columns=["Артикул WB","Списание за отзывы"])
    total_other=0.0
    if "deduction" in df_raw.columns and "bonus_type_name" in df_raw.columns:
        mask_rev=df_raw["bonus_type_name"].str.contains("списание за отзыв",case=False,na=False)
        revs=df_raw[mask_rev&(df_raw["deduction"]!=0)].assign(**{
            "Артикул WB": lambda d: d["bonus_type_name"].str.extract(r"товар\s+(\d+)")[0].str.upper()
        })
        if not revs.empty:
            reviews_agg=revs.groupby("Артикул WB",as_index=False)["deduction"].sum().rename(columns={"deduction":"Списание за отзывы"})
        mask_other=(df_raw["deduction"]!=0)&~df_raw["bonus_type_name"].str.contains(
            "подписке «Джем»|Списание за отзыв|ВБ\.?Продвижение|Акт утилизации товара",
            case=False,na=False
        )
        total_other += df_raw.loc[mask_other,"deduction"].sum()
    if "penalty" in df_raw.columns:
        total_penalty = df_raw.loc[
            (df_raw["nm_id"] == 0) & (df_raw["penalty"] != 0),
            "penalty"
        ].sum()
        print(total_penalty)
        total_other += total_penalty

    # объединяем
    for df,col in [(sales_df,"Артикул WB"),(storage_df,"nmId"),(adv_df,"Артикул WB"),(acceptance_df,"Артикул WB")]:
        df[col]=df[col].astype(str).str.upper()

    merged=pd.merge(sales_df, storage_df.rename(columns={"nmId":"Артикул WB"})[["Артикул WB","vendorCode","totalStorageSum"]],on="Артикул WB",how="outer")
    merged=pd.merge(merged, adv_df[["Артикул WB","totalAdjustedSum"]],on="Артикул WB",how="outer")
    merged=pd.merge(merged, acceptance_df[["Артикул WB","Платная приемка"]],on="Артикул WB",how="outer")
    merged=pd.merge(merged, reviews_agg, on="Артикул WB",how="left").fillna(0)
    merged.fillna(0,inplace=True)
    merged.sort_values("Артикул WB",inplace=True)

    # Прочие удержания
    n=len(merged)
    per_item=round(total_other/n,2) if n else 0.0
    merged["Прочие удержания"]=per_item

    # Переименование и порядок
    merged.rename(columns={
        "Короткое название товара":"Артикул поставщика",
        "SUM из Кол-во":"Кол-во продаж",
        "SUM из Сумма продаж":"Общая выручка",
        "SUM из К перечислению продавцу":"К Перечислению",
        "SUM из Кол-во доставок":"Логистика, шт",
        "SUM из Стоимость доставки":"Логистика, руб",
        "SUM из Штрафы":"Штрафы",
        "SUM из Дополнительный платеж":"Доплаты",
        "Возвраты (К перечислению продавцу)":"Возвраты",
        "totalStorageSum":"Хранение",
        "totalAdjustedSum":"ВБ.Продвижение"
    },inplace=True)

    final_cols=[
        "Артикул WB","Артикул поставщика","Кол-во продаж","Общая выручка",
        "К Перечислению","Логистика, шт","Логистика, руб","Штрафы","Доплаты",
        "Возвраты","Хранение","ВБ.Продвижение","Подписка «Джем»",
        "Платная приемка","Утилизация","Списание за отзывы","Прочие удержания"
    ]
    final_df=merged[final_cols]

    # Добавляем столбец "На расчетный счет"
    final_df["На расчетный счет"] = (
        final_df["К Перечислению"]
        - final_df["Логистика, руб"]
        - final_df["Штрафы"]
        + final_df["Доплаты"]
        - final_df["Возвраты"]
        - final_df["Хранение"]
        - final_df["ВБ.Продвижение"]
        - final_df["Подписка «Джем»"]
        - final_df["Платная приемка"]
        - final_df["Утилизация"]
        - final_df["Списание за отзывы"]
        - final_df["Прочие удержания"]
    )

    yellow=PatternFill(fill_type="solid",start_color="FFFF00",end_color="FFFF00")

    output_folder = Path('data') / 'reports' / str(tg_id) / str(store_id)  # /data on server
    output_folder.mkdir(parents=True, exist_ok=True)
    path = output_folder / f'report{start_date}.xlsx'

    with pd.ExcelWriter(path,engine="openpyxl") as writer:
        final_df.to_excel(writer,index=False,startrow=2)
        ws=writer.sheets["Sheet1"]
        ws.cell(row=1,column=1,value=f"Магазин: {store_name}")
        ws.cell(row=2,column=1,value=f"Период: {start_date} – {end_date}")
        for cell in ws[3]:
            cell.font=Font(bold=True)
        for col in ws.columns:
            length=max(len(str(c.value)) for c in col)
            ws.column_dimensions[col[0].column_letter].width=length+2
        summary=ws.max_row+1
        ws.cell(row=summary,column=1,value="Итого").font=Font(bold=True)
        for idx in range(3,len(final_cols)+2):
            letter=get_column_letter(idx)
            c=ws.cell(row=summary,column=idx,value=f"=SUM({letter}4:{letter}{summary-1})")
            c.font=Font(color="FF0000"); c.fill=yellow

    logger.info(f'Итоговый отчёт сохранён в "{path}"')
    return str(path)